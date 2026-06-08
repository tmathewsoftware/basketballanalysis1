import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from team_matcher import find_team_id

HIGHLIGHTLY_API_KEY = "5a5ad2b7-dc79-4187-af30-418b7bd28cae"
BASE_URL = "https://basketball.highlightly.net"
HEADERS = {"x-rapidapi-key": HIGHLIGHTLY_API_KEY}

# ── Shared Styles ─────────────────────────────────────────────────────────────
thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center")
GREEN_FILL = PatternFill("solid", start_color="70AD47")
FLOOR_FILL = PatternFill("solid", start_color="D9D9D9")
HALF_FILL = PatternFill("solid", start_color="E8E8E8")
HEADER_FILL = PatternFill("solid", start_color="1A1A2E")
TITLE_FILL = PatternFill("solid", start_color="0D0D3E")

LOWER_IS_BETTER = {"Turnovers", "Personal Fouls", "Flagrant Fouls", "Technical Fouls"}

STAT_ORDER = [
    "Succesful Field Goals", "Field Goals",
    "Succesful 3 Pointers", "3 Pointers",
    "Succesful Free Throws", "Free Throws",
    "Assists", "Rebounds", "Offensive Rebounds", "Defensive Rebounds",
    "Steals", "Blocks", "Turnovers",
    "Fast Break Points", "Points Off Turnovers", "Points In The Paint",
    "Personal Fouls", "Second Chance Points", "Biggest Lead",
    "Flagrant Fouls", "Technical Fouls"
]


# ── Helpers ───────────────────────────────────────────────────────────────────
def write_cell(ws, row, col, value, font=None, fill=None, bold=False, is_half=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font or Font(name="Arial", size=10, bold=bold)
    cell.alignment = CENTER
    cell.border = BORDER
    if fill:
        cell.fill = fill
    elif is_half:
        cell.fill = HALF_FILL
    return cell


def parse_score(score_str):
    try:
        parts = score_str.split(" - ")
        return int(parts[0]), int(parts[1])
    except:
        return 0, 0


def parse_quarters(score_dict, home_is_team1):
    quarters = []
    for q in ["q1", "q2", "q3", "q4"]:
        val = score_dict.get(q, "0 - 0") or "0 - 0"
        home_q, away_q = parse_score(val)
        if home_is_team1:
            quarters.append((home_q, away_q))
        else:
            quarters.append((away_q, home_q))
    return quarters


# ── API Calls ─────────────────────────────────────────────────────────────────
def get_h2h_games(team1_id, team2_id):
    url = f"{BASE_URL}/head-2-head"
    response = requests.get(url, headers=HEADERS, params={
        "teamIdOne": team1_id,
        "teamIdTwo": team2_id
    })
    data = response.json()
    if isinstance(data, list):
        return data
    return data.get("data", [])


def get_match_statistics(match_id):
    url = f"{BASE_URL}/matches/{match_id}"
    response = requests.get(url, headers=HEADERS)
    data = response.json()
    if isinstance(data, list) and len(data) > 0:
        return data[0].get("statistics", [])
    return []


# ── Table 1: H2H Game Totals ──────────────────────────────────────────────────
def write_h2h_totals(ws, games, team1_id, team1_name, team2_name, start_row=1):
    ws.merge_cells(f"A{start_row}:D{start_row}")
    c = ws.cell(row=start_row, column=1, value=f"H2H Game Totals — {team1_name} vs {team2_name}")
    c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill = TITLE_FILL
    c.alignment = CENTER
    ws.row_dimensions[start_row].height = 28

    for col, header in enumerate(["Date", team1_name, team2_name, "Combined"], 1):
        cell = ws.cell(row=start_row + 1, column=col, value=header)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[start_row + 1].height = 22

    t1_scores, t2_scores, combined_scores = [], [], []

    for i, game in enumerate(games):
        row = start_row + 2 + i
        try:
            date_str = datetime.strptime(game["date"][:10], "%Y-%m-%d").strftime("%d %b %Y")
        except:
            date_str = game.get("date", "")[:10]

        score_str = game.get("state", {}).get("score", {}).get("current", "0 - 0")
        home_score, away_score = parse_score(score_str)
        home_id = game.get("homeTeam", {}).get("id")
        t1, t2 = (home_score, away_score) if home_id == team1_id else (away_score, home_score)
        combined = t1 + t2
        t1_scores.append(t1)
        t2_scores.append(t2)
        combined_scores.append(combined)

        write_cell(ws, row, 1, date_str)
        c1 = write_cell(ws, row, 2, t1)
        c2 = write_cell(ws, row, 3, t2)
        write_cell(ws, row, 4, combined)

        if t1 > t2:
            c1.fill = GREEN_FILL
            c1.font = Font(name="Arial", size=10, bold=True)
        elif t2 > t1:
            c2.fill = GREEN_FILL
            c2.font = Font(name="Arial", size=10, bold=True)

        ws.row_dimensions[row].height = 20

    fc_row = start_row + 2 + len(games)
    write_cell(ws, fc_row, 1, "Floor / Ceiling", bold=True, fill=FLOOR_FILL)
    for col, scores in enumerate([t1_scores, t2_scores, combined_scores], 2):
        val = f"{min(scores)}-{max(scores)}" if scores else "N/A"
        write_cell(ws, fc_row, col, val, bold=True, fill=FLOOR_FILL)
    ws.row_dimensions[fc_row].height = 20

    return fc_row


# ── Table 2: Quarter/Half Breakdown (side by side) ────────────────────────────
def write_quarter_tables(ws, games, team1_id, team1_name, team2_name, start_row=1):
    t1_col = 1
    t2_col = 8

    for team_col, team_name, is_team1 in [
        (t1_col, team1_name, True),
        (t2_col, team2_name, False)
    ]:
        ws.merge_cells(
            start_row=start_row, start_column=team_col,
            end_row=start_row, end_column=team_col + 6
        )
        c = ws.cell(row=start_row, column=team_col, value=f"{team_name} — Quarter / Half Breakdown")
        c.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
        c.fill = TITLE_FILL
        c.alignment = CENTER
        ws.row_dimensions[start_row].height = 28

        for j, h in enumerate(["Date", "Q1", "Q2", "Q3", "Q4", "H1", "H2"]):
            cell = ws.cell(row=start_row + 1, column=team_col + j, value=h)
            cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = BORDER
        ws.row_dimensions[start_row + 1].height = 22

        q_scores = [[], [], [], [], [], []]

        for i, game in enumerate(games):
            row = start_row + 2 + i
            try:
                date_str = datetime.strptime(game["date"][:10], "%Y-%m-%d").strftime("%d %b %Y")
            except:
                date_str = game.get("date", "")[:10]

            score_dict = game.get("state", {}).get("score", {})
            home_id = game.get("homeTeam", {}).get("id")
            home_is_team1 = (home_id == team1_id)
            quarters = parse_quarters(score_dict, home_is_team1)

            if is_team1:
                q_vals = [q[0] for q in quarters]
                opp_vals = [q[1] for q in quarters]
            else:
                q_vals = [q[1] for q in quarters]
                opp_vals = [q[0] for q in quarters]

            h1 = q_vals[0] + q_vals[1]
            h2 = q_vals[2] + q_vals[3]
            opp_h1 = opp_vals[0] + opp_vals[1]
            opp_h2 = opp_vals[2] + opp_vals[3]
            all_vals = q_vals + [h1, h2]
            all_opp = opp_vals + [opp_h1, opp_h2]

            for idx, v in enumerate(all_vals):
                q_scores[idx].append(v)

            write_cell(ws, row, team_col, date_str)
            for j, (val, opp) in enumerate(zip(all_vals, all_opp)):
                col = team_col + 1 + j
                is_half = j >= 4
                cell = write_cell(ws, row, col, val, is_half=is_half)
                if val > opp:
                    cell.fill = GREEN_FILL
                    cell.font = Font(name="Arial", size=10, bold=True)
            ws.row_dimensions[row].height = 20

        fc_row = start_row + 2 + len(games)
        write_cell(ws, fc_row, team_col, "Floor / Ceiling", bold=True, fill=FLOOR_FILL)
        for j, scores in enumerate(q_scores):
            val = f"{min(scores)}-{max(scores)}" if scores else "N/A"
            write_cell(ws, fc_row, team_col + 1 + j, val, bold=True, fill=FLOOR_FILL)
        ws.row_dimensions[fc_row].height = 20

    return fc_row


# ── Table 3: Combined Quarter/Half Totals ─────────────────────────────────────
def write_combined_quarter_table(ws, games, team1_id, team1_name, team2_name, start_row=1):
    ws.merge_cells(
        start_row=start_row, start_column=1,
        end_row=start_row, end_column=7
    )
    c = ws.cell(row=start_row, column=1, value=f"Combined Quarter / Half Totals — {team1_name} vs {team2_name}")
    c.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    c.fill = TITLE_FILL
    c.alignment = CENTER
    ws.row_dimensions[start_row].height = 28

    for j, h in enumerate(["Date", "Q1", "Q2", "Q3", "Q4", "H1", "H2"]):
        cell = ws.cell(row=start_row + 1, column=1 + j, value=h)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[start_row + 1].height = 22

    q_scores = [[], [], [], [], [], []]

    for i, game in enumerate(games):
        row = start_row + 2 + i
        try:
            date_str = datetime.strptime(game["date"][:10], "%Y-%m-%d").strftime("%d %b %Y")
        except:
            date_str = game.get("date", "")[:10]

        score_dict = game.get("state", {}).get("score", {})
        home_id = game.get("homeTeam", {}).get("id")
        home_is_team1 = (home_id == team1_id)
        quarters = parse_quarters(score_dict, home_is_team1)

        q_vals = [q[0] + q[1] for q in quarters]
        h1 = q_vals[0] + q_vals[1]
        h2 = q_vals[2] + q_vals[3]
        all_vals = q_vals + [h1, h2]

        for idx, v in enumerate(all_vals):
            q_scores[idx].append(v)

        write_cell(ws, row, 1, date_str)
        for j, val in enumerate(all_vals):
            write_cell(ws, row, 2 + j, val, is_half=(j >= 4))
        ws.row_dimensions[row].height = 20

    fc_row = start_row + 2 + len(games)
    write_cell(ws, fc_row, 1, "Floor / Ceiling", bold=True, fill=FLOOR_FILL)
    for j, scores in enumerate(q_scores):
        val = f"{min(scores)}-{max(scores)}" if scores else "N/A"
        write_cell(ws, fc_row, 2 + j, val, bold=True, fill=FLOOR_FILL)
    ws.row_dimensions[fc_row].height = 20

    return fc_row


# ── Table 4: Game Statistics ──────────────────────────────────────────────────
def write_stats_table(ws, games, team1_id, team2_id, team1_name, team2_name, start_row=1):
    num_games = len(games)
    total_cols = 1 + num_games * 2

    # Title
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=total_cols)
    c = ws.cell(row=start_row, column=1, value=f"Game Statistics — {team1_name} vs {team2_name}")
    c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill = TITLE_FILL
    c.alignment = CENTER
    ws.row_dimensions[start_row].height = 28

    # Stat label header
    cell = ws.cell(row=start_row + 1, column=1, value="Stat")
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER

    # Date headers — merged across 2 cols per game
    for g_idx, game in enumerate(games):
        try:
            date_str = datetime.strptime(game["date"][:10], "%Y-%m-%d").strftime("%d %b %Y")
        except:
            date_str = game.get("date", "")[:10]
        col_start = 2 + g_idx * 2
        ws.merge_cells(start_row=start_row + 1, start_column=col_start, end_row=start_row + 1, end_column=col_start + 1)
        c = ws.cell(row=start_row + 1, column=col_start, value=date_str)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[start_row + 1].height = 22

    # Team name sub-headers
    ws.cell(row=start_row + 2, column=1).border = BORDER
    for g_idx in range(num_games):
        col_start = 2 + g_idx * 2
        for offset, name in enumerate([team1_name, team2_name]):
            cell = ws.cell(row=start_row + 2, column=col_start + offset, value=name)
            cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = BORDER
    ws.row_dimensions[start_row + 2].height = 18

    # Fetch stats for each game
    print("Fetching game statistics...")
    all_stats = []
    for game in games:
        match_id = game.get("id")
        stats_data = get_match_statistics(match_id)
        team_stats = {}
        for team_entry in stats_data:
            tid = team_entry.get("team", {}).get("id")
            stat_dict = {s["displayName"]: s["value"] for s in team_entry.get("statistics", [])}
            team_stats[tid] = stat_dict
        all_stats.append(team_stats)

    # Write stat rows
    for s_idx, stat_name in enumerate(STAT_ORDER):
        row = start_row + 3 + s_idx

        cell = ws.cell(row=row, column=1, value=stat_name)
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.alignment = CENTER
        cell.border = BORDER
        ws.row_dimensions[row].height = 18

        for g_idx in range(num_games):
            team_stats = all_stats[g_idx]
            t1_stats = team_stats.get(team1_id, {})
            t2_stats = team_stats.get(team2_id, {})
            t1_val = t1_stats.get(stat_name, "-")
            t2_val = t2_stats.get(stat_name, "-")

            col1 = 2 + g_idx * 2
            col2 = col1 + 1

            c1 = ws.cell(row=row, column=col1, value=t1_val)
            c2 = ws.cell(row=row, column=col2, value=t2_val)

            for c in [c1, c2]:
                c.font = Font(name="Arial", size=10)
                c.alignment = CENTER
                c.border = BORDER

            if isinstance(t1_val, (int, float)) and isinstance(t2_val, (int, float)):
                lower_better = stat_name in LOWER_IS_BETTER
                if lower_better:
                    if t1_val < t2_val:
                        c1.fill = GREEN_FILL
                        c1.font = Font(name="Arial", size=10, bold=True)
                    elif t2_val < t1_val:
                        c2.fill = GREEN_FILL
                        c2.font = Font(name="Arial", size=10, bold=True)
                else:
                    if t1_val > t2_val:
                        c1.fill = GREEN_FILL
                        c1.font = Font(name="Arial", size=10, bold=True)
                    elif t2_val > t1_val:
                        c2.fill = GREEN_FILL
                        c2.font = Font(name="Arial", size=10, bold=True)

    last_row = start_row + 3 + len(STAT_ORDER) - 1

    # Column widths for stats table
    ws.column_dimensions["A"].width = 24
    for g_idx in range(num_games):
        for offset in range(2):
            col_idx = 2 + g_idx * 2 + offset
            ws.column_dimensions[get_column_letter(col_idx)].width = 14

    return last_row


# ── Main Builder ──────────────────────────────────────────────────────────────
def build_h2h_excel(team1_name, team2_name, output_path="h2h_report.xlsx"):
    print("Finding team IDs...")
    team1_id, team1_matched, _ = find_team_id(team1_name)
    team2_id, team2_matched, _ = find_team_id(team2_name)

    if not team1_id or not team2_id:
        print("Could not find one or both teams. Aborting.")
        return

    print("Fetching H2H games...")
    games = get_h2h_games(team1_id, team2_id)

    if not games:
        print("No H2H games found.")
        return

    games = games[:5]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "H2H Analysis"

    # Set default column widths
    for col_idx in range(1, 16):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    # Table 1: H2H Totals
    last_row = write_h2h_totals(ws, games, team1_id, team1_matched, team2_matched, start_row=1)

    # Table 2: Quarter/Half breakdown
    quarter_start = last_row + 2
    last_row = write_quarter_tables(ws, games, team1_id, team1_matched, team2_matched, start_row=quarter_start)

    # Table 3: Combined quarter/half totals
    combined_start = last_row + 2
    last_row = write_combined_quarter_table(ws, games, team1_id, team1_matched, team2_matched, start_row=combined_start)

    # Table 4: Game statistics
    stats_start = last_row + 2
    write_stats_table(ws, games, team1_id, team2_id, team1_matched, team2_matched, start_row=stats_start)

    wb.save(output_path)
    print(f"Saved to {output_path}")


if __name__ == "__main__":
    team1 = "Los Angeles Lakers"
    team2 = "San Antonio Spurs"
    date_str = datetime.now().strftime("%d %b %Y")
    filename = f"H2H — {team1} vs {team2} — {date_str}.xlsx"
    build_h2h_excel(team1, team2, filename)