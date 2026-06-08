import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from team_matcher import find_team_id

HIGHLIGHTLY_API_KEY = "5a5ad2b7-dc79-4187-af30-418b7bd28cae"
BASE_URL = "https://basketball.highlightly.net"
HEADERS = {"x-rapidapi-key": HIGHLIGHTLY_API_KEY}

thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center")
GREEN_FILL = PatternFill("solid", start_color="70AD47")
FLOOR_FILL = PatternFill("solid", start_color="D9D9D9")
HALF_FILL = PatternFill("solid", start_color="E8E8E8")
HEADER_FILL = PatternFill("solid", start_color="1A1A2E")
TITLE_FILL = PatternFill("solid", start_color="0D0D3E")

LOWER_IS_BETTER = {"Turnovers", "Personal Fouls", "Flagrant Fouls", "Technical Fouls"}

STAT_ORDER = [
    "Succesful Field Goals", "Field Goals",
    "Succesful 3 Pointers", "3 Pointers",
    "Succesful Free Throws", "Free Throws",
    "Assists", "Rebounds", "Offensive Rebounds", "Defensive Rebounds",
    "Steals", "Blocks", "Turnovers",
    "Fast Break Points", "Points Off Turnovers", "Points In The Paint",
    "Personal Fouls", "Second Chance Points", "Biggest Lead",
    "Flagrant Fouls", "Technical Fouls"
]


def write_cell(ws, row, col, value, font=None, fill=None, bold=False, is_half=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font or Font(name="Arial", size=10, bold=bold)
    cell.alignment = CENTER
    cell.border = BORDER
    if fill:
        cell.fill = fill
    elif is_half:
        cell.fill = HALF_FILL
    return cell


def parse_score(score_str):
    try:
        parts = score_str.split(" - ")
        return int(parts[0]), int(parts[1])
    except:
        return 0, 0


def parse_quarters(score_dict, team_is_home):
    quarters = []
    for q in ["q1", "q2", "q3", "q4"]:
        val = score_dict.get(q, "0 - 0") or "0 - 0"
        home_q, away_q = parse_score(val)
        if team_is_home:
            quarters.append((home_q, away_q))
        else:
            quarters.append((away_q, home_q))
    return quarters


def get_opponent_name(game, team_id):
    home_id = game.get("homeTeam", {}).get("id")
    if home_id == team_id:
        return game.get("awayTeam", {}).get("name", "Unknown")
    return game.get("homeTeam", {}).get("name", "Unknown")


def get_last_five_games(team_id):
    url = f"{BASE_URL}/last-five-games"
    response = requests.get(url, headers=HEADERS, params={"teamId": team_id})
    data = response.json()
    if isinstance(data, list):
        return data
    return data.get("data", [])


def get_match_statistics(match_id):
    url = f"{BASE_URL}/matches/{match_id}"
    response = requests.get(url, headers=HEADERS)
    data = response.json()
    if isinstance(data, list) and len(data) > 0:
        return data[0].get("statistics", [])
    return []


def write_game_totals(ws, games, team_id, team_name, start_row=1, start_col=1):
    end_col = start_col + 4

    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
    c = ws.cell(row=start_row, column=start_col, value=f"{team_name} — Recent Game Totals")
    c.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    c.fill = TITLE_FILL
    c.alignment = CENTER
    ws.row_dimensions[start_row].height = 28

    for j, h in enumerate(["Date", "Opponent", team_name, "Opp Score", "Combined"]):
        cell = ws.cell(row=start_row + 1, column=start_col + j, value=h)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[start_row + 1].height = 22

    team_scores, opp_scores, combined_scores = [], [], []

    for i, game in enumerate(games):
        row = start_row + 2 + i
        try:
            date_str = datetime.strptime(game["date"][:10], "%Y-%m-%d").strftime("%d %b %Y")
        except:
            date_str = game.get("date", "")[:10]

        score_str = game.get("state", {}).get("score", {}).get("current", "0 - 0")
        home_score, away_score = parse_score(score_str)
        home_id = game.get("homeTeam", {}).get("id")
        team_is_home = (home_id == team_id)
        t_score = home_score if team_is_home else away_score
        o_score = away_score if team_is_home else home_score
        combined = t_score + o_score
        opp_name = get_opponent_name(game, team_id)

        team_scores.append(t_score)
        opp_scores.append(o_score)
        combined_scores.append(combined)

        write_cell(ws, row, start_col, date_str)
        write_cell(ws, row, start_col + 1, opp_name)
        c1 = write_cell(ws, row, start_col + 2, t_score)
        c2 = write_cell(ws, row, start_col + 3, o_score)
        write_cell(ws, row, start_col + 4, combined)

        if t_score > o_score:
            c1.fill = GREEN_FILL
            c1.font = Font(name="Arial", size=10, bold=True)
        elif o_score > t_score:
            c2.fill = GREEN_FILL
            c2.font = Font(name="Arial", size=10, bold=True)

        ws.row_dimensions[row].height = 20

    fc_row = start_row + 2 + len(games)
    write_cell(ws, fc_row, start_col, "Floor / Ceiling", bold=True, fill=FLOOR_FILL)
    write_cell(ws, fc_row, start_col + 1, "", fill=FLOOR_FILL)
    for j, scores in enumerate([team_scores, opp_scores, combined_scores]):
        val = f"{min(scores)}-{max(scores)}" if scores else "N/A"
        write_cell(ws, fc_row, start_col + 2 + j, val, bold=True, fill=FLOOR_FILL)
    ws.row_dimensions[fc_row].height = 20

    return fc_row


def write_quarter_breakdown(ws, games, team_id, team_name, start_row=1, start_col=1):
    end_col = start_col + 7

    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
    c = ws.cell(row=start_row, column=start_col, value=f"{team_name} — Quarter / Half Breakdown")
    c.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    c.fill = TITLE_FILL
    c.alignment = CENTER
    ws.row_dimensions[start_row].height = 28

    for j, h in enumerate(["Date", "Opponent", "Q1", "Q2", "Q3", "Q4", "H1", "H2"]):
        cell = ws.cell(row=start_row + 1, column=start_col + j, value=h)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[start_row + 1].height = 22

    q_scores = [[], [], [], [], [], []]

    for i, game in enumerate(games):
        row = start_row + 2 + i
        try:
            date_str = datetime.strptime(game["date"][:10], "%Y-%m-%d").strftime("%d %b %Y")
        except:
            date_str = game.get("date", "")[:10]

        score_dict = game.get("state", {}).get("score", {})
        home_id = game.get("homeTeam", {}).get("id")
        team_is_home = (home_id == team_id)
        quarters = parse_quarters(score_dict, team_is_home)
        opp_name = get_opponent_name(game, team_id)

        q_vals = [q[0] for q in quarters]
        opp_vals = [q[1] for q in quarters]
        h1 = q_vals[0] + q_vals[1]
        h2 = q_vals[2] + q_vals[3]
        opp_h1 = opp_vals[0] + opp_vals[1]
        opp_h2 = opp_vals[2] + opp_vals[3]
        all_vals = q_vals + [h1, h2]
        all_opp = opp_vals + [opp_h1, opp_h2]

        for idx, v in enumerate(all_vals):
            q_scores[idx].append(v)

        write_cell(ws, row, start_col, date_str)
        write_cell(ws, row, start_col + 1, opp_name)
        for j, (val, opp) in enumerate(zip(all_vals, all_opp)):
            is_half = j >= 4
            cell = write_cell(ws, row, start_col + 2 + j, val, is_half=is_half)
            if val > opp:
                cell.fill = GREEN_FILL
                cell.font = Font(name="Arial", size=10, bold=True)
        ws.row_dimensions[row].height = 20

    fc_row = start_row + 2 + len(games)
    write_cell(ws, fc_row, start_col, "Floor / Ceiling", bold=True, fill=FLOOR_FILL)
    write_cell(ws, fc_row, start_col + 1, "", fill=FLOOR_FILL)
    for j, scores in enumerate(q_scores):
        val = f"{min(scores)}-{max(scores)}" if scores else "N/A"
        write_cell(ws, fc_row, start_col + 2 + j, val, bold=True, fill=FLOOR_FILL)
    ws.row_dimensions[fc_row].height = 20

    return fc_row


def write_stats_table(ws, games, team_id, team_name, start_row=1):
    num_games = len(games)
    total_cols = 1 + num_games * 2

    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=total_cols)
    c = ws.cell(row=start_row, column=1, value=f"{team_name} — Game Statistics")
    c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill = TITLE_FILL
    c.alignment = CENTER
    ws.row_dimensions[start_row].height = 28

    cell = ws.cell(row=start_row + 1, column=1, value="Stat")
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER

    for g_idx, game in enumerate(games):
        try:
            date_str = datetime.strptime(game["date"][:10], "%Y-%m-%d").strftime("%d %b %Y")
        except:
            date_str = game.get("date", "")[:10]
        col_start = 2 + g_idx * 2
        ws.merge_cells(start_row=start_row + 1, start_column=col_start, end_row=start_row + 1, end_column=col_start + 1)
        c = ws.cell(row=start_row + 1, column=col_start, value=date_str)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[start_row + 1].height = 22

    ws.cell(row=start_row + 2, column=1).border = BORDER
    for g_idx, game in enumerate(games):
        opp_name = get_opponent_name(game, team_id)
        col_start = 2 + g_idx * 2
        for offset, name in enumerate([team_name, opp_name]):
            cell = ws.cell(row=start_row + 2, column=col_start + offset, value=name)
            cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = BORDER
    ws.row_dimensions[start_row + 2].height = 18

    print(f"  Fetching stats for {team_name}...")
    all_stats = []
    for game in games:
        match_id = game.get("id")
        stats_data = get_match_statistics(match_id)
        team_stats = {}
        for team_entry in stats_data:
            tid = team_entry.get("team", {}).get("id")
            stat_dict = {s["displayName"]: s["value"] for s in team_entry.get("statistics", [])}
            team_stats[tid] = stat_dict
        all_stats.append(team_stats)

    for s_idx, stat_name in enumerate(STAT_ORDER):
        row = start_row + 3 + s_idx
        cell = ws.cell(row=row, column=1, value=stat_name)
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.alignment = CENTER
        cell.border = BORDER
        ws.row_dimensions[row].height = 18

        for g_idx, game in enumerate(games):
            home_id = game.get("homeTeam", {}).get("id")
            opp_id = game.get("awayTeam", {}).get("id") if home_id == team_id else home_id
            team_stats = all_stats[g_idx]
            t_stats = team_stats.get(team_id, {})
            o_stats = team_stats.get(opp_id, {})
            t_val = t_stats.get(stat_name, "-")
            o_val = o_stats.get(stat_name, "-")

            col1 = 2 + g_idx * 2
            col2 = col1 + 1
            c1 = ws.cell(row=row, column=col1, value=t_val)
            c2 = ws.cell(row=row, column=col2, value=o_val)

            for c in [c1, c2]:
                c.font = Font(name="Arial", size=10)
                c.alignment = CENTER
                c.border = BORDER

            if isinstance(t_val, (int, float)) and isinstance(o_val, (int, float)):
                lower_better = stat_name in LOWER_IS_BETTER
                if lower_better:
                    if t_val < o_val:
                        c1.fill = GREEN_FILL
                        c1.font = Font(name="Arial", size=10, bold=True)
                    elif o_val < t_val:
                        c2.fill = GREEN_FILL
                        c2.font = Font(name="Arial", size=10, bold=True)
                else:
                    if t_val > o_val:
                        c1.fill = GREEN_FILL
                        c1.font = Font(name="Arial", size=10, bold=True)
                    elif o_val > t_val:
                        c2.fill = GREEN_FILL
                        c2.font = Font(name="Arial", size=10, bold=True)

    last_row = start_row + 3 + len(STAT_ORDER) - 1
    ws.column_dimensions["A"].width = 24
    for g_idx in range(num_games):
        for offset in range(2):
            col_idx = 2 + g_idx * 2 + offset
            ws.column_dimensions[get_column_letter(col_idx)].width = 14

    return last_row


def build_recent_games_excel(team1_name, team2_name, output_path=None):
    print("Finding team IDs...")
    team1_id, team1_matched, _ = find_team_id(team1_name)
    team2_id, team2_matched, _ = find_team_id(team2_name)

    if not team1_id or not team2_id:
        print("Could not find one or both teams. Aborting.")
        return

    print("Fetching last 5 games...")
    t1_games = get_last_five_games(team1_id)
    t2_games = get_last_five_games(team2_id)

    if not t1_games or not t2_games:
        print("Could not fetch recent games for one or both teams.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Recent Games"

    for col_idx in range(1, 25):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    # Section 1: Game Totals side by side (5 cols each, 1 gap)
    last_row_t1 = write_game_totals(ws, t1_games, team1_id, team1_matched, start_row=1, start_col=1)
    write_game_totals(ws, t2_games, team2_id, team2_matched, start_row=1, start_col=7)

    section2_start = last_row_t1 + 2

    # Section 2: Quarter/Half Breakdown side by side (8 cols each, 1 gap)
    last_row_q = write_quarter_breakdown(ws, t1_games, team1_id, team1_matched, start_row=section2_start, start_col=1)
    write_quarter_breakdown(ws, t2_games, team2_id, team2_matched, start_row=section2_start, start_col=10)

    section3_start = last_row_q + 2

    # Section 3: Stats tables stacked vertically
    last_row_s1 = write_stats_table(ws, t1_games, team1_id, team1_matched, start_row=section3_start)
    write_stats_table(ws, t2_games, team2_id, team2_matched, start_row=last_row_s1 + 2)

    if not output_path:
        date_str = datetime.now().strftime("%d %b %Y")
        output_path = f"Recent Games — {team1_matched} vs {team2_matched} — {date_str}.xlsx"

    wb.save(output_path)
    print(f"Saved to {output_path}")


if __name__ == "__main__":
    build_recent_games_excel("Los Angeles Lakers", "San Antonio Spurs")